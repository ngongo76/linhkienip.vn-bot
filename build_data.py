"""
Build Vietnamese Excel + JSON from the original Bot_Fixmobile.xlsx
Output:
  - Bot_Fixmobile_VN.xlsx (for Kevin to review/edit)
  - panic_data.json (for the bot to read)
"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================
# DATA: All entries with Vietnamese translations
# ============================================================
# Format: (code, meaning_en, meaning_vi)
# For i2c: (code, cpu_series, chip_vi)
# For Bảng chữ cái: (letter, code, meaning_en, meaning_vi)

SENSOR_ARRAY_1 = [
    ("0X400",   "Try to Reball Interposer",                                          "Đóng lại chip Interposer (Reball)"),
    ("0X800",   "It's the Charging Port Flex",                                       "Cáp chân sạc"),
    ("0X1000",  "It's the Proximity Flex Cable",                                     "Cáp cảm biến tiệm cận (Prox Flex)"),
    ("0X1800",  "It's Both: Charging Port Flex & Prox Flex",                         "Cả 2: Cáp chân sạc + Cáp tiệm cận"),
    ("0X4000",  "It's part of the Battery",                                          "Liên quan đến pin"),
    ("0X10000", "It's the Power Button Flex",                                        "Cáp nút nguồn"),
    ("0X20000", "It's a Sandwich Board Issue / Gyro",                                "Lỗi bo sandwich / Cảm biến con quay (Gyro)"),
    ("0X40000", "It's the Charging Port Flex",                                       "Cáp chân sạc"),
    ("0X60000", "It's the Proximity Flex Cable",                                     "Cáp cảm biến tiệm cận"),
    ("0X80000", "It's the Proximity Flex Cable",                                     "Cáp cảm biến tiệm cận"),
    ("0X140000","It's the Charging Port & Power Button Flex or Barometer",          "Cáp chân sạc + Cáp nút nguồn HOẶC cảm biến khí áp"),
    ("0X180000","It's the Prox Flex & Power Button Flex",                            "Cáp tiệm cận + Cáp nút nguồn"),
    ("0X41",    "It's a Battery Data issue",                                         "Lỗi dữ liệu pin (Battery Data)"),
    ("0X42",    "Indicates a thermal issue, such as overheating",                    "Lỗi nhiệt độ / Máy quá nhiệt"),
    ("0X51",    "Indicates a problem with the device's security features or encryption", "Lỗi bảo mật hoặc mã hóa của máy"),
    ("0X61",    "Indicates a problem with the device's display, such as flickering or pixelation", "Lỗi màn hình: nhấp nháy hoặc vỡ điểm ảnh"),
    ("0X63",    "Indicates a problem with the device's accelerometer or motion sensor", "Lỗi cảm biến gia tốc / cảm biến chuyển động"),
    ("0X64",    "Indicates a problem with the device's microphone or audio input",   "Lỗi mic / cổng thu âm"),
    ("0X71",    "Indicates a problem with the device's audio or speaker system",     "Lỗi hệ thống âm thanh / loa"),
    ("0X73",    "Indicates a problem with the device's ambient light sensor or automatic brightness adjustment", "Lỗi cảm biến ánh sáng môi trường / tự động chỉnh độ sáng"),
    ("0X74",    "Indicates a problem with the device's speaker or audio output",     "Lỗi loa / ngõ ra âm thanh"),
    ("0X81",    "Indicates a problem with the device's camera or imaging system",    "Lỗi camera / hệ thống chụp ảnh"),
    ("0X83",    "Indicates a problem with the device's proximity sensor or screen auto-dimming feature", "Lỗi cảm biến tiệm cận / tự tắt màn khi áp tai"),
    ("0X84",    "Indicates a problem with the device's vibration motor or haptic feedback", "Lỗi mô tơ rung / Taptic Engine"),
    ("0X91",    "Indicates a problem with the device's network connectivity, such as Wi-Fi or cellular signal issues", "Lỗi kết nối mạng: WiFi hoặc sóng di động"),
    ("0X93",    "Indicates a problem with the device's gyroscope",                   "Lỗi con quay hồi chuyển (Gyro)"),
    ("0X94",    "Indicates a problem with the device's camera or imaging system",    "Lỗi camera / hệ thống chụp ảnh"),
    ("0X104",   "Indicates a problem with the device's accelerometer or gravity sensor", "Lỗi cảm biến gia tốc / cảm biến trọng lực"),
    ("0X114",   "Indicates a problem with the device's gyroscope or rotation sensor", "Lỗi con quay / cảm biến xoay"),
    ("0X124",   "Indicates a problem with the device's magnetometer or compass sensor", "Lỗi cảm biến từ trường / la bàn"),
    ("0X134",   "Indicates a problem with the device's proximity sensor or light sensor", "Lỗi cảm biến tiệm cận / cảm biến ánh sáng"),
    ("0X144",   "Indicates a problem with the device's ambient temperature sensor",  "Lỗi cảm biến nhiệt độ môi trường"),
    ("0X154",   "Indicates a problem with the device's humidity sensor",             "Lỗi cảm biến độ ẩm"),
    ("0X164",   "Indicates a problem with the device's pressure sensor",             "Lỗi cảm biến áp suất"),
    ("0X174",   "Indicates a problem with the device's accelerometer or gravity sensor", "Lỗi cảm biến gia tốc / trọng lực"),
    ("0X184",   "Indicates a problem with the device's gyroscope or rotation sensor", "Lỗi con quay / cảm biến xoay"),
    ("0X194",   "Indicates a problem with the device's magnetometer or compass sensor", "Lỗi cảm biến từ trường / la bàn"),
    ("0X204",   "Indicates a problem with the device's humidity sensor or moisture detection system", "Lỗi cảm biến độ ẩm / phát hiện nước"),
    ("0X214",   "Indicates a problem with the device's pressure sensor or air pressure detection mechanism", "Lỗi cảm biến áp suất / cơ chế đo áp suất không khí"),
]

SENSOR_ARRAY_2 = [
    ("0X1A4",   "Indicates a problem with the device's proximity sensor or light sensor", "Lỗi cảm biến tiệm cận / cảm biến ánh sáng"),
    ("0X1B4",   "Indicates a problem with the device's ambient light sensor or automatic brightness adjustment", "Lỗi cảm biến ánh sáng môi trường / tự chỉnh độ sáng"),
    ("0X1C4",   "Indicates a problem with the device's proximity sensor or screen auto-dimming feature", "Lỗi cảm biến tiệm cận / tự tắt màn khi áp tai"),
    ("0X1D4",   "Indicates a problem with the device's gyroscope or orientation sensor", "Lỗi con quay / cảm biến phương hướng"),
    ("0X1E4",   "Indicates a problem with the device's barometer or altitude sensor", "Lỗi cảm biến khí áp (Barometer) / cảm biến độ cao"),
    ("0X1F4",   "Indicates a problem with the device's thermometer or temperature sensor", "Lỗi cảm biến nhiệt độ"),
    ("0XD4",    "Indicates a problem with the device's ambient temperature sensor", "Lỗi cảm biến nhiệt độ môi trường"),
    ("0XE4",    "Indicates a problem with the device's humidity sensor",            "Lỗi cảm biến độ ẩm"),
    ("0XF4",    "Indicates a problem with the device's pressure sensor",            "Lỗi cảm biến áp suất"),
    ("0XA4",    "Indicates a problem with the device's touch screen or digitizer",  "Lỗi cảm ứng / digitizer"),
    ("0XB4",    "Indicates a problem with the device's SIM card or cellular connectivity", "Lỗi khe SIM / kết nối di động"),
    ("0XC4",    "Indicates a problem with the device's fingerprint sensor or biometric authentication", "Lỗi cảm biến vân tay / xác thực sinh trắc"),
    ("0XD1",    "Indicates a problem with the device's vibration motor or haptic feedback", "Lỗi mô tơ rung / Taptic Engine"),
    ("0XE1",    "Indicates a problem with the device's ambient light sensor or proximity sensor", "Lỗi cảm biến ánh sáng môi trường / cảm biến tiệm cận"),
    ("0XF1",    "Indicates a problem with the device's gyroscope or accelerometer", "Lỗi con quay / cảm biến gia tốc"),
    ("0XA1",    "Indicates a problem with the device's GPS or location services",   "Lỗi GPS / dịch vụ định vị"),
    ("0XB1",    "Indicates a problem with the device's Touch ID or Face ID functionality", "Lỗi Touch ID / Face ID"),
    ("0XC1",    "Indicates a problem with the device's microphone or audio input",  "Lỗi mic / cổng thu âm"),
    ("0Xa1",    "Battery Communications",                                            "Giao tiếp pin (Battery Communications)"),
    ("0XC0000", "It's the Prox Flex & Charging Port Flex",                           "Cáp tiệm cận + Cáp chân sạc"),
    ("0X100000","It's the Power Button Flex / Charging Port Flex",                   "Cáp nút nguồn / Cáp chân sạc"),
    ("0X1C0000","It's the Charging Port & Power Button Flex & Prox Flex",            "Cáp chân sạc + Cáp nút nguồn + Cáp tiệm cận"),
    ("0X200000","It's the Proximity Flex Cable",                                     "Cáp cảm biến tiệm cận"),
    ("0X280000","Charging flex & wireless charging flex",                            "Cáp chân sạc + Cáp sạc không dây"),
    ("0X300000","It's the Proximity Flex Cable / Indicates camera or imaging system issues", "Cáp tiệm cận / Lỗi camera"),
    ("0X400000","Wireless Charging Flex (back glass)",                               "Cáp sạc không dây (mặt lưng kính)"),
    ("0X500000","Battery Communications / NFC",                                      "Giao tiếp pin / NFC"),
    ("0X600000","Wireless charging flex & proximity flex",                           "Cáp sạc không dây + Cáp tiệm cận"),
    ("0X700000","Charging flex & wireless charging flex",                            "Cáp chân sạc + Cáp sạc không dây"),
    ("0X800000","Software-related issues",                                           "Lỗi phần mềm"),
    ("0XA00000","Battery-related issues",                                            "Lỗi liên quan đến pin"),
    ("0XB00000","Audio-related issues (speakers, microphones, headphone jack)",      "Lỗi âm thanh (loa, mic, jack tai nghe)"),
    ("0XC00000","Security features (Touch ID, Face ID, or Secure Enclave)",          "Bảo mật (Touch ID, Face ID, hoặc Secure Enclave)"),
    ("0XD00000","Sensor issues (accelerometer, gyroscope, proximity, or ambient light)", "Lỗi cảm biến (gia tốc, con quay, tiệm cận, ánh sáng)"),
    ("0XF00000","Miscellaneous hardware issues (buttons, connectors, overall integrity)", "Lỗi phần cứng tổng hợp (nút, đầu nối, kết cấu tổng thể)"),
]

USERSPACE_WATCHDOG = [
    ("TT1P / TT2P",                "Screen",                                "Màn hình"),
    ("TG0B",                       "Battery No Data",                       "Pin không có dữ liệu"),
    ("TP3R",                       "NTC Problem",                           "Lỗi NTC (cảm biến nhiệt pin)"),
    ("TP1A",                       "Battery Failure",                       "Pin hỏng"),
    ("TP2A / TP3R / TP4H",         "Battery/Screen",                        "Pin / Màn hình"),
    ("TP2C",                       "Battery/Screen",                        "Pin / Màn hình"),
    ("TG0V / TTSA",                "Charging Port Flex",                    "Cáp chân sạc"),
    ("Prs0",                       "Charging Port Flex / U7400",            "Cáp chân sạc / IC U7400"),
    ("MIC1",                       "Charging Port Flex",                    "Cáp chân sạc"),
    ("MIC2",                       "Power Key Flex",                        "Cáp nút nguồn"),
    ("SpringBoard",                "Restore",                               "Restore (khôi phục lại iOS)"),
    ("600seconds",                 "Hard Disk (Nand)",                      "Bộ nhớ (Nand)"),
    ("Without the above code",     "U2 Priority (TRISTAR-HYDRA)",           "Ưu tiên kiểm tra U2 (Tristar/Hydra - IC sạc)"),
]

I2C = [
    ("i2c0",     "A8 Series (iPhone 6 / 6 Plus)",         "U1202, U1501, U1502, U1700"),
    ("i2c0",     "A9 Series (iPhone 6s / 6s Plus)",       "U2000, U4000, U4020"),
    ("i2c0",     "A10 Series (iPhone 7 / 7 Plus)",        "U1801, U3703, U3701, U4001, U2301"),
    ("i2c0",     "A11 Series (iPhone 8 / 8 Plus / X)",    "U2700, U5600, U5660, U6110, J6400"),
    ("i2c0",     "A12 Series (iPhone Xs / Xs Max / XR)",  "U2700, U6110, J6400"),
    ("i2c1",     "A8 Series",                             "U1580, U1400, U1401, U1601, J2118"),
    ("i2c1",     "A9 Series",                             "U3800, U2300, U3700, U4500"),
    ("i2c1",     "A10 Series",                            "U1801, U2101, U4601"),
    ("i2c1",     "A11 Series",                            "J4300, J6400"),
    ("i2c1",     "A12 Series",                            "J4300"),
    ("i2c2",     "A8 Series",                             "J1111, J2019"),
    ("i2c2",     "A9 Series",                             "J3100, J4200, U4050"),
    ("i2c2",     "A10 Series",                            "U3301, J4503"),
    ("i2c2",     "A11 Series",                            "U3301, J4200, U5000"),
    ("i2c2",     "A12 Series",                            "U5002"),
    ("i2c3",     "A10 Series",                            "Chân tail, Mic sau hoặc Màn hình"),
    ("i2c3",     "A11 Series (iPhone 8 / 8 Plus)",        "Chân tail, Socket màn, Màn, IC đèn nền (U5650)"),
    ("i2c3",     "A11 Series (iPhone X)",                 "Nguồn màn, Cảm ứng, Đế cảm ứng"),
    ("i2c3",     "A12 Series",                            "Nguồn màn, Cảm ứng, Đế cảm ứng"),
    ("i2c4",     "A11 Series",                            "Chip logic"),
    ("i2c5",     "A10 Series",                            "Chip logic"),
    ("SMC i2c0", "A11 Series",                            "U3100 / U3300 / U3400 / U6200 / J3200"),
    ("SMC i2c0", "A12 Series",                            "U3300 / U3400 / U6200 / J3200"),
    ("SMC i2c1", "A11 Series",                            "IC nguồn chính USB"),
    ("SMC i2c1", "A12 Series",                            "IC nguồn chính USB"),
]

AOP_PANIC = [
    ("SCMto",                            "Vibration Motor",                                                                                "Mô tơ rung"),
    ("Systick watchdog",                 "Big Audio IC",                                                                                   "IC âm thanh lớn (Audio bự)"),
    ("Systick watchdog2 not pet",        "Gyroscope",                                                                                      "Con quay hồi chuyển (Gyro)"),
    ("No pulse on",                      "Big Audio IC / Front Speaker / Vibration",                                                       "IC âm thanh lớn / Loa thoại / Mô tơ rung"),
    ("AMCC ERROR",                       "Front Speaker",                                                                                  "Loa thoại"),
    ("AOP PANIC (7P)",                   "Proximity Sensor / Rear MIC (iPhone 7 Plus) / Power Button Flex / Fingerprint (Touch ID)",       "Cảm biến tiệm cận / Mic sau (7 Plus) / Cáp nút nguồn / Vân tay (Touch ID)"),
    ("( AOP PANIC )",                    "Big Audio IC / Screen Failure",                                                                  "IC âm thanh lớn / Lỗi màn hình"),
    ("AOP DATA ABORT",                   "Hard Drive (Nand Flash) / CPU Solder Joint (Reball CPU)",                                        "Bộ nhớ (Nand) / Mối hàn CPU (cần đóng lại chip CPU)"),
    ("AOP data link CMD timeout;bus 0",  "AOP data link CMD timeout;bus 0",                                                                "Lỗi timeout đường dữ liệu AOP - bus 0"),
    ("K2-bosch control",                 "Vibrator (Taptic Engine)",                                                                       "Mô tơ rung (Taptic Engine)"),
    ("PressureController.cpp:280",       "Layer inspection L3302 / Rear Camera / Distance Sensor / Rear MIC (7P Boot Cable) / Damaged Fingerprint Flex", "Kiểm tra lớp main L3302 / Camera sau / Cảm biến khoảng cách / Mic sau (Cáp boot 7P) / Cáp vân tay hỏng"),
]

ALPHABET = [
    ("A", "Attemping to forcibly halt CPU",                                  "CPU Solder Joint (Requires CPU Reballing)",                                            "Mối hàn CPU (cần đóng lại chip CPU - Reball)"),
    ("A", "AMCC Error",                                                      "Light-Sensitive Sensor / Ambient Light Sensor",                                        "Cảm biến ánh sáng / Cảm biến ánh sáng môi trường"),
    ("A", "AppleBCMWLAN",                                                    "WiFi / Bluetooth Module",                                                              "Module WiFi / Bluetooth"),
    ("A", "Anc-postnand.c1260 asser failed link",                            "Nand Problem (Hard Disk)",                                                             "Lỗi bộ nhớ Nand"),
    ("A", "AGXK AGXAcceletor",                                               "AppleSOC Overheating / Gyroscope / Accelerator / Coprocessor / WiFi / Big Audio IC / CPU Reballing", "Quá nhiệt CPU Apple / Con quay / Cảm biến gia tốc / Đồng xử lý / WiFi / IC âm thanh lớn / Reball CPU"),
    ("A", "ANS",                                                             "Hard Disk Circuit or Board Layer",                                                     "Mạch bộ nhớ hoặc lớp main"),
    ("A", "ANS2",                                                            "Hard Disk Circuit or Board Layer",                                                     "Mạch bộ nhớ hoặc lớp main"),
    ("A", "AOP NMI POWER",                                                   "Front Flex Cable / Power Button Cable",                                                "Cáp mặt trước / Cáp nút nguồn"),
    ("A", "Apple PMGR",                                                      "Battery",                                                                              "Pin"),
    ("A", "Apple PPM",                                                       "Charging IC / Battery Conversion Circuit",                                             "IC sạc / Mạch chuyển đổi pin"),
    ("A", "apcie (wlan)",                                                    "WiFi Module",                                                                          "Module WiFi"),
    ("A", "A freed zone element has been modified",                          "CPU Solder Joint (Requires CPU Reballing)",                                            "Mối hàn CPU (cần Reball CPU)"),
    ("A", "A kext releasing",                                                "First consider the CPU Inductance",                                                    "Ưu tiên kiểm tra cuộn cảm CPU"),
    ("A", "apcie(0:s3e)",                                                    "Hard Disk (Nand)",                                                                     "Bộ nhớ (Nand)"),
    ("A", "AppleHIDTransportProtocol",                                       "Screen / Display Circuit",                                                             "Màn hình / Mạch màn hình"),
    ("A", "APFS_TREE",                                                       "Hard Disk (Nand)",                                                                     "Bộ nhớ (Nand)"),
    ("B", "Busy timeout",                                                    "Pmic / Audio",                                                                         "IC nguồn (PMIC) / Âm thanh"),
    ("B", "Bad tallq elm",                                                   "Main Crystal Oscillator / Wide-Angle Camera / Screen",                                 "Thạch anh chính / Camera góc rộng / Màn hình"),
    ("B", "BASEBAND",                                                        "Baseband Part",                                                                        "Khối baseband (modem sóng)"),
    ("C", "cluster_push err:expectedNULL",                                   "Load SoftWare",                                                                        "Nạp lại phần mềm"),
    ("C", "Coherency point error",                                           "Check The Rear Camera",                                                                "Kiểm tra camera sau"),
    ("C", "CP_COM_NORM REQUEST",                                             "CPU Hard Disk / Camera Circuit",                                                       "Khối CPU - bộ nhớ / Mạch camera"),
    ("C", "CSL AUDIO",                                                       "Audio Related",                                                                        "Liên quan đến âm thanh"),
    ("C", "CAN Snoop",                                                       "Ply Stub",                                                                             "Mỏ hàn / Đầu hàn lớp main"),
    ("D", "DCP PANIC",                                                       "Charging flex",                                                                        "Cáp chân sạc"),
    ("D", "DPC err UNKNOWN:multiple) on E-core",                             "PMU Buck, Voltage or CPU",                                                             "PMU Buck, điện áp hoặc CPU"),
    ("D", "Dart-disp0 SMMU error",                                           "Rear Camera",                                                                          "Camera sau"),
    ("D", "DPC err",                                                         "Hard disk and baseband circuit, crystal oscillator",                                   "Mạch bộ nhớ + mạch baseband, thạch anh"),
    ("E", "Ememory",                                                         "Hard Disk or Hard Disk Related Circuit",                                               "Bộ nhớ hoặc mạch liên quan đến bộ nhớ"),
    ("E", "enableHalogen:3833 Unknow sample rate on DAC",                    "Audio",                                                                                "Âm thanh"),
    ("F", "Firmware fatal",                                                  "Firmware or Underlying or Power Button Line",                                          "Firmware / Lớp dưới / Đường nút nguồn"),
    ("F", "fed err(parity counter overflow multi hit ICTAG)",                "Hard Disk",                                                                            "Bộ nhớ"),
    ("F", "Fatal coherency point error CP_com_NORM",                         "BUCK CPU Power Supply Problem, CPU Side Inductance",                                   "Lỗi nguồn BUCK cấp cho CPU, cuộn cảm cạnh CPU"),
    ("H", "Halt\\Restart Timed out",                                         "Tail Plug / Earphone Amplifier",                                                       "Chân tail / IC khuếch đại tai nghe"),
    ("I", "Iokit",                                                           "Iokit",                                                                                "Iokit (driver hệ thống)"),
    ("I", "Initproc exited -- exit",                                         "Try Software or CPU Reball",                                                           "Thử phần mềm hoặc Reball CPU"),
    ("I", "Invaild queue element linkage",                                   "The Hard Disk is Welded or Damaged",                                                   "Bộ nhớ bị bong chân hoặc hỏng"),
    ("I", "Initproc exited",                                                 "Main crystal failure",                                                                 "Thạch anh chính hỏng"),
    ("I", "IOMFB int hander",                                                "Screen or screen 0.8V power supply 16",                                                "Màn hình hoặc nguồn 0.8V cấp màn"),
    ("K", "Key hdr kind!=NEW",                                               "Display IC",                                                                           "IC màn hình"),
    ("K", "Kernel data abort",                                               "Restore / Hard Disk / Board Layer",                                                    "Restore / Bộ nhớ / Lớp main"),
    ("L", "L2C",                                                             "WIFI / Audio Amplification",                                                           "WiFi / Khuếch đại âm thanh"),
    ("L", "LLC Bus error from cpu3 : FAR",                                   "Abnormal Communication between CPU and WiFi",                                          "Giao tiếp bất thường giữa CPU và WiFi"),
    ("L", "LSU",                                                             "Wifi",                                                                                 "WiFi"),
    ("L", "LLC",                                                             "Baseband Circuitry Possibly Audio",                                                    "Mạch baseband, có thể là âm thanh"),
    ("L", "lop_rinobuffer",                                                  "Audio Related",                                                                        "Liên quan đến âm thanh"),
    ("M", "Menorystaus_jetsam_thread",                                       "Hard Disk",                                                                            "Bộ nhớ"),
    ("M", "Mbuf_watchdog:12 waiters stuck",                                  "Boot Cable / Tail Plug / Power Button",                                                "Cáp boot / Chân tail / Nút nguồn"),
    ("M", "mubf_watchdog:2waiters stuck",                                    "Hard Disk has bad sectors",                                                            "Bộ nhớ bị bad sector"),
    ("N", "Nvme",                                                            "Hard Disk",                                                                            "Bộ nhớ"),
    ("N", "NO pulse on",                                                     "Vibration / Pulse (ring vibration)",                                                   "Mô tơ rung / Tín hiệu rung khi có chuông"),
    ("P", "Pmap_enter_pv. unexpected PV head",                               "Earpiece Row, Small Audio",                                                            "Hàng loa thoại, âm thanh nhỏ"),
    ("P", "PMP NMI FIQ",                                                     "CPU Power Supply",                                                                     "Nguồn cấp cho CPU"),
    ("P", "Pre, next",                                                       "Main Crystal Failure",                                                                 "Thạch anh chính hỏng"),
    ("R", "Reset sequence did not finish within S000ms",                     "Abnormal Communication between CPU and WiFi",                                          "Giao tiếp bất thường giữa CPU và WiFi"),
    ("R", "Rtunref Bad refcnt",                                              "Hard Disk",                                                                            "Bộ nhớ"),
    ("R", "RTBuddy ( AOP ) setmanagedGated",                                 "Battery Holder",                                                                       "Khay pin"),
    ("R", "Releasing non_exclusive",                                         "CPU Inductor",                                                                         "Cuộn cảm CPU"),
    ("S", "Sep memory protection module error",                              "EEPROM (Screen / CPU / Baseband)",                                                     "EEPROM (Màn / CPU / Baseband)"),
    ("S", "SEP ROM",                                                         "FingerPrint ROM / Logic Chip",                                                         "ROM vân tay / Chip logic"),
    ("S", "Sep ROM boot Panic [Scrubbed",                                    "EEPROM CPU, Resistors, or CPU",                                                        "EEPROM của CPU, điện trở hoặc CPU"),
    ("S", "SEP Panic: :dxio/ I2C : 0X100007785",                             "EEPROM CPU, Resistors, or CPU",                                                        "EEPROM của CPU, điện trở hoặc CPU"),
    ("S", "SEP Panic: :sars/ sars: 0X10000f439",                             "EEPROM CPU, Resistors, or CPU",                                                        "EEPROM của CPU, điện trở hoặc CPU"),
    ("S", "SMC DATA ABORT",                                                  "Abnormal CPU Communication",                                                           "Giao tiếp CPU bất thường"),
    ("S", "Stacks+routined-2019 01 29",                                      "The iPad is mostly caused by the Battery Screw post being too long (Special iPad issue)", "iPad: thường do ốc vít gắn pin quá dài (Lỗi đặc trưng iPad)"),
    ("S", "sks request timeout",                                             "CPU to Logic Code Fragment Line / EEPROM",                                             "Đường code logic từ CPU / EEPROM"),
    ("S", "Sleep \\wake hang detected",                                      "Battery, Audio CPU Powered",                                                           "Pin, nguồn cấp CPU âm thanh"),
    ("S", "SEP panic:: daio/i2c",                                            "Down Antenna Problem",                                                                 "Lỗi anten dưới"),
    ("S", "SMC PANIC",                                                       "CPU Power Supply Check Inductance",                                                    "Nguồn CPU - kiểm tra cuộn cảm"),
    ("S", "Spmi timeout",                                                    "Audio Related",                                                                        "Liên quan đến âm thanh"),
    ("S", "SEP Panic sars/sars",                                             "Earpiece Row, Small Audio",                                                            "Hàng loa thoại, âm thanh nhỏ"),
    ("T", "Timeout waiting for stop to IDLE on EP",                          "Screen",                                                                               "Màn hình"),
    ("T", "Trying to change a collection in the registry",                   "Just Flash",                                                                           "Chỉ cần nạp lại phần mềm"),
    ("V", "Void applesynopsysMIPID SIController",                            "Front Cable",                                                                          "Cáp trước"),
    ("W", "WDT timeout",                                                     "Battery or Battery Connector",                                                         "Pin hoặc connector pin"),
    ("W", "WKDMD ERROR code 0x2",                                            "Flashing error 14 / Hard Disk (Nand Flash)",                                           "Lỗi flash 14 / Bộ nhớ (Nand)"),
]

# ============================================================
# BUILD EXCEL FILE
# ============================================================

PURPLE_FILL = PatternFill('solid', start_color='5D52B7')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
BODY_FONT = Font(name='Arial', size=10)
CODE_FONT = Font(name='Consolas', size=10, bold=True, color='5D52B7')
THIN = Side(border_style='thin', color='CCCCCC')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center')

def style_header(cell):
    cell.fill = PURPLE_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    cell.border = BORDER

def style_body(cell, code=False):
    cell.font = CODE_FONT if code else BODY_FONT
    cell.alignment = WRAP
    cell.border = BORDER

def add_sheet_3col(wb, name, headers, rows):
    ws = wb.create_sheet(name)
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        style_header(c)
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            style_body(cell, code=(c == 1))
    widths = [22, 55, 55]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.row_dimensions[1].height = 28
    for r in range(2, len(rows) + 2):
        ws.row_dimensions[r].height = 42
    ws.freeze_panes = 'A2'

def add_sheet_4col(wb, name, headers, rows):
    ws = wb.create_sheet(name)
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        style_header(c)
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            style_body(cell, code=(c == 2))
    widths = [8, 40, 45, 45]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.row_dimensions[1].height = 28
    for r in range(2, len(rows) + 2):
        ws.row_dimensions[r].height = 42
    ws.freeze_panes = 'A2'

wb = Workbook()
wb.remove(wb.active)

add_sheet_3col(wb, "Sensor Array (1)",
               ["Mã lỗi", "Ý nghĩa (EN)", "Ý nghĩa (Tiếng Việt)"],
               SENSOR_ARRAY_1)
add_sheet_3col(wb, "Sensor Array (2)",
               ["Mã lỗi", "Ý nghĩa (EN)", "Ý nghĩa (Tiếng Việt)"],
               SENSOR_ARRAY_2)
add_sheet_3col(wb, "Userspace watchdog",
               ["Mã lỗi", "Ý nghĩa (EN)", "Ý nghĩa (Tiếng Việt)"],
               USERSPACE_WATCHDOG)
add_sheet_3col(wb, "i2c",
               ["Mã lỗi", "Dòng CPU / Model", "Chip / Linh kiện liên quan"],
               I2C)
add_sheet_3col(wb, "AOP PANIC",
               ["Mã lỗi", "Ý nghĩa (EN)", "Ý nghĩa (Tiếng Việt)"],
               AOP_PANIC)
add_sheet_4col(wb, "Bảng chữ cái",
               ["Chữ", "Mã lỗi", "Giải thích (EN)", "Giải thích (Tiếng Việt)"],
               ALPHABET)

wb.save("/home/claude/fixmobile_bot/Bot_Linhkienip_VN.xlsx")
print("✓ Excel saved")

# ============================================================
# BUILD JSON FILE
# ============================================================

data = {
    "categories": {
        "sensor_1": {
            "title": "Sensor Array (1)",
            "title_vi": "Mã cảm biến nhóm 1",
            "entries": [{"code": c, "en": e, "vi": v} for c, e, v in SENSOR_ARRAY_1],
        },
        "sensor_2": {
            "title": "Sensor Array (2)",
            "title_vi": "Mã cảm biến nhóm 2",
            "entries": [{"code": c, "en": e, "vi": v} for c, e, v in SENSOR_ARRAY_2],
        },
        "watchdog": {
            "title": "Userspace watchdog timeout",
            "title_vi": "Lỗi watchdog (treo tiến trình)",
            "entries": [{"code": c, "en": e, "vi": v} for c, e, v in USERSPACE_WATCHDOG],
        },
        "i2c": {
            "title": "i2c Bus",
            "title_vi": "Lỗi bus i2c (giao tiếp giữa chip)",
            "entries": [{"code": c, "cpu": cpu, "vi": chip} for c, cpu, chip in I2C],
        },
        "aop": {
            "title": "AOP PANIC",
            "title_vi": "Lỗi AOP (vi xử lý luôn bật)",
            "entries": [{"code": c, "en": e, "vi": v} for c, e, v in AOP_PANIC],
        },
    },
    "alphabet": {},
}

for letter, code, en, vi in ALPHABET:
    data["alphabet"].setdefault(letter, []).append({"code": code, "en": en, "vi": vi})

with open("/home/claude/fixmobile_bot/panic_data.json", "w", encoding="utf-8") as f:
    json.dump(data, f, ensure_ascii=False, indent=2)
print("✓ JSON saved")

# Summary
total = (len(SENSOR_ARRAY_1) + len(SENSOR_ARRAY_2) + len(USERSPACE_WATCHDOG)
         + len(I2C) + len(AOP_PANIC) + len(ALPHABET))
print(f"Total entries: {total}")
print(f"  Sensor Array 1: {len(SENSOR_ARRAY_1)}")
print(f"  Sensor Array 2: {len(SENSOR_ARRAY_2)}")
print(f"  Watchdog: {len(USERSPACE_WATCHDOG)}")
print(f"  i2c: {len(I2C)}")
print(f"  AOP PANIC: {len(AOP_PANIC)}")
print(f"  Alphabet: {len(ALPHABET)}")
